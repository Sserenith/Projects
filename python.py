import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

"Load in worksheet"
wb = openpyxl.load_workbook(Bay Area raw.xlsx)

"creates new sheets for removed and processed tables"
wb.create_sheet(title=REED processed)
wb.create_sheet(title=removed)
sheet = wb.worksheets[0]
sheet.title = Old
scrubsheet = wb[REED processed]
removedsheet = wb[removed]
print(n)

"open keyword files"
with open('subtype.txt', 'r') as f
    content_s = f.read()
subtype_exceptions = content_s.split(n)
with open('names.txt', 'r') as g
    content_n = g.read()
name_exceptions = content_n.split(n)

"column and row counters"
rowscrub = scrubsheet.max_row + 1
rowremoved = removedsheet.max_row + 1
scrubbedcount = 0
removedcount = 0
column_count = sheet.max_column

"copies the header of the table to the new sheets"
for h in range(1, column_count + 1)
    source = sheet.cell(row=1, column=h)
    dest = removedsheet.cell(row=1, column=h)
    dest.value = source.value
    dest = scrubsheet.cell(row=1, column=h)
    dest.value = source.value

"""used try - except in order to get past an openpyxl bug that causes the table to not have a max row value.
following script parses through columns, comparing cell values to keywords in txt files, and moving rows to removed if they contain the keywords""
or moves them to scrubsheet processed if they do not in order to maintain the original file unmodified."""
try
    for t in range(sheet.min_row+1, sheet.max_row)
        if any([x in sheet.cell(row=t, column=10).value for x in subtype_exceptions]) or any([y in sheet.cell(row=t, column=3).value for y in name_exceptions])
            for c in range(1, column_count + 1)
                hl_obj = sheet.cell(row=t, column=3).hyperlink
                if hl_obj
                    source = sheet.cell(row=t, column=3)
                    dest = removedsheet.cell(row=rowremoved, column=3)
                    dest.value = source.value
                    dest.hyperlink = source.hyperlink
                    dest.style = Hyperlink
                source = sheet.cell(row=t, column=c)
                dest = removedsheet.cell(row=rowremoved, column=c)
                dest.value = source.value
            rowremoved += 1
            removedcount += 1
        else
            for c1 in range(1, column_count + 1)
                hl_obj = sheet.cell(row=t, column=3).hyperlink
                if hl_obj
                    source = sheet.cell(row=t, column=3)
                    dest = scrubsheet.cell(row=rowscrub, column=3)
                    dest.value = source.value
                    dest.hyperlink = source.hyperlink
                    dest.style = Hyperlink
                source = sheet.cell(row=t, column=c1)
                dest = scrubsheet.cell(row=rowscrub, column=c1)
                dest.value = source.value
            rowscrub += 1
            scrubbedcount += 1
except
    print(Extra rows, stopped the for loop)

"prints a summary of the task"
print(f"Of {scrubbedcount+removedcount} projects, {scrubbedcount} are viable and {removedcount} have been removed")

"removes the unneeded worksheets from the excel file. Can be commented for testing"
wb.remove(wb[removed])
wb.remove(wb[Old])

"table formatting"
scrubsheettab = Table(displayName=REED_processed, ref=fA1P{rowscrub-1})
style = TableStyleInfo(name=TableStyleMedium2, showRowStripes=True)
scrubsheettab.tableStyleInfo = style
scrubsheet.add_table(scrubsheettab)
removedsheettab = Table(displayName=removed, ref=fA1P{rowscrub-1})
style = TableStyleInfo(name=TableStyleLight1, showRowStripes=True)
removedsheettab.tableStyleInfo = style
removedsheet.add_table(removedsheettab)

"saves the result in a new worksheet"
wb.save(Bay Area raw _pyscrubbed.xlsx)